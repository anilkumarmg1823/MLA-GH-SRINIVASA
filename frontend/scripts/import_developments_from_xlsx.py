# -*- coding: utf-8 -*-
"""Import development works from constituency kamagari xlsx into JSON + JS seed."""
from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
BACKEND = ROOT / "backend"
GP_JS = ROOT / "frontend" / "src" / "data" / "gramPanchayats.js"
OUT_JSON = ROOT / "backend" / "data" / "developmentsFromXlsx.json"
OUT_JS = ROOT / "frontend" / "src" / "data" / "developments.js"
REPORT = ROOT / "frontend" / "scripts" / "_dev_import_report.json"

# Kannada literals via unicode escapes (keeps this file ASCII-safe)
def u(s: str) -> str:
    return s.encode("utf-8").decode("unicode_escape") if "\\u" in s else s


SKIP_SHEETS = {
    "Sheet1",
    "Sheet2",
    "Sheet3",
    "Sheet4",
    "Sheet5",
    u("\\u0c92\\u0c9f\\u0ccd\\u0c9f\\u0cc1 \\u0c95\\u0cbe\\u0cae\\u0c97\\u0cbe\\u0cb0\\u0cbf\\u0c97\\u0cb3 \\u0cb5\\u0cbf\\u0cb5\\u0cb0"),
}

SHEET_GP_MAP = {
    u("\\u0c85\\u0caa\\u0ccd\\u0caa\\u0cc7\\u0ca8\\u0cb9\\u0cb3\\u0ccd\\u0cb3\\u0cbf"): "Appenahalli",
    u("\\u0c97\\u0cc1\\u0ca1\\u0cc7\\u0c95\\u0ccb\\u0c9f\\u0cc6"): "Gudekote",
    u("\\u0cb0\\u0cbe\\u0cae\\u0ca6\\u0cc1\\u0cb0\\u0ccd\\u0c97"): "Ramdurga",
    u("\\u0cac\\u0cc6\\u0cb3\\u0ccd\\u0cb3\\u0c97\\u0c9f\\u0ccd\\u0c9f"): "Bellagatta",
    u("\\u0c9c\\u0cb0\\u0ccd\\u0cae\\u0cb2\\u0cbf"): "Jarmali",
    u("\\u0c97\\u0cc1\\u0c82\\u0ca1\\u0cc1\\u0cae\\u0cc1\\u0ca3\\u0cc1\\u0c97\\u0cc1"): "Gundumunugu",
    u("\\u0c9a\\u0cbf\\u0cb0\\u0ca4\\u0c97\\u0cc1\\u0c82\\u0ca1\\u0cc1"): "Chirathagundu",
    u("\\u0cb9\\u0cc1\\u0cb0\\u0cc1\\u0cb3\\u0cbf\\u0cb9\\u0cbe\\u0cb3\\u0cc1"): "Huralihal",
    u("\\u0cae\\u0cbe\\u0c95\\u0ca8\\u0ca1\\u0c95\\u0cc1"): "Makanadaku",
    u("\\u0c97\\u0c82\\u0ca1\\u0cac\\u0cca\\u0cae\\u0ccd\\u0cae\\u0ca8\\u0cb9\\u0cb3\\u0ccd\\u0cb3\\u0cbf"): "Gambommanahalli",
    u("\\u0cb9\\u0cc1\\u0ca1\\u0cc7\\u0c82"): "Hudem",
    u("\\u0c9c\\u0cc1\\u0cae\\u0ccd\\u0cae\\u0cca\\u0cac\\u0ca8\\u0cb9\\u0cb3\\u0ccd\\u0cb3\\u0cbf"): "Jummobanahalli",
    u("\\u0caa\\u0cc2\\u0c9c\\u0cbe\\u0cb0\\u0cb9\\u0cb3\\u0ccd\\u0cb3\\u0cbf"): "Pujarahalli",
    u("\\u0c86\\u0cb2\\u0cc2\\u0cb0\\u0cc1"): "Alur",
    u("\\u0cb9\\u0cbf\\u0cb0\\u0cc7\\u0c95\\u0cc1\\u0c82\\u0cac\\u0cb3\\u0c97\\u0cc1\\u0c82\\u0c9f\\u0cc6"): "Hirekumbalgunte",
    u("\\u0cb9\\u0cca\\u0cb8\\u0cb9\\u0cb3\\u0ccd\\u0cb3\\u0cbf"): "Hosahalli",
    u("\\u0cac\\u0ca3\\u0cb5\\u0cbf\\u0c95\\u0cb2\\u0ccd\\u0cb2\\u0cc1"): "Banavikallu",
    u("\\u0cb9\\u0cbe\\u0cb0\\u0c95\\u0cac\\u0cbe\\u0cb5\\u0cbf"): "Harakbavi",
    u("\\u0cb8\\u0cc2\\u0cb2\\u0ca6\\u0cb9\\u0cb3\\u0ccd\\u0cb3\\u0cbf"): "Suladahalli",
    u("\\u0c9a\\u0ccc\\u0ca1\\u0cbe\\u0caa\\u0cc1\\u0cb0"): "Chowdapur",
    u("\\u0cac\\u0ca1\\u0cc7\\u0cb2\\u0ca1\\u0c95\\u0cc1"): "Badeladaku",
    u("\\u0cb9\\u0cbf\\u0cb0\\u0cc7\\u0cb9\\u0cc6\\u0c97\\u0ccd\\u0ca1\\u0cbe\\u0cb3\\u0ccd"): "Hirehegdal",
    u("\\u0cb6\\u0cbf\\u0cb5\\u0caa\\u0cc1\\u0cb0"): "Shivpur",
    u("\\u0cae\\u0cca\\u0cb0\\u0cac"): "Moraba",
    u("\\u0c95\\u0c95\\u0ccd\\u0c95\\u0cc1\\u0caa\\u0ccd\\u0caa\\u0cbf"): "Kakkuppi",
    u("\\u0c89\\u0c9c\\u0ccd\\u0c9c\\u0cbf\\u0ca8\\u0cbf"): "Ujjini",
    u("\\u0ca8\\u0cbf\\u0c82\\u0cac\\u0cb3\\u0c97\\u0cc6\\u0cb0\\u0cc6"): "Nimbalagere",
    u("\\u0ca4\\u0cc2\\u0cb2\\u0cb9\\u0cb3\\u0ccd\\u0cb3\\u0cbf"): "Tulahalli",
    u("\\u0c95\\u0cbe\\u0cb3\\u0cbe\\u0caa\\u0cc1\\u0cb0"): "Kalapur",
    u("\\u0ca8\\u0cbe\\u0c97\\u0cb0\\u0c95\\u0c9f\\u0ccd\\u0c9f\\u0cc6"): "Nagarkatte",
    u("\\u0cb8\\u0cc1\\u0c82\\u0c95\\u0ca6\\u0c95\\u0cb2\\u0ccd\\u0cb2\\u0cc1"): "Sunkadakallu",
    u("\\u0c95\\u0cc2\\u0ca1\\u0ccd\\u0cb2\\u0cbf\\u0c97\\u0cbf \\u0caa\\u0c9f\\u0ccd\\u0c9f\\u0ca3"): "Kudligi Town",
    u("\\u0c95\\u0cc2\\u0ca1\\u0ccd\\u0cb2\\u0cbf\\u0c97\\u0cbf \\u0cb5\\u0cbf\\u0ca7\\u0cbe\\u0ca8\\u0cb8\\u0cad\\u0cbe"): "Kudligi Constituency",
}

SUNKA_KN = u("\\u0cb8\\u0cc1\\u0c82\\u0c95\\u0ca6\\u0c95\\u0cb2\\u0ccd\\u0cb2\\u0cc1")
KUD_TOWN_KN = u("\\u0c95\\u0cc2\\u0ca1\\u0ccd\\u0cb2\\u0cbf\\u0c97\\u0cbf \\u0caa\\u0c9f\\u0ccd\\u0c9f\\u0ca3")
KUD_CONST_KN = u("\\u0c95\\u0cc2\\u0ca1\\u0ccd\\u0cb2\\u0cbf\\u0c97\\u0cbf \\u0cb5\\u0cbf\\u0ca7\\u0cbe\\u0ca8\\u0cb8\\u0cad\\u0cbe")
KUD_KN = u("\\u0c95\\u0cc2\\u0ca1\\u0ccd\\u0cb2\\u0cbf\\u0c97\\u0cbf")
TALUK_KN = u("\\u0c95\\u0cc2\\u0ca1\\u0ccd\\u0cb2\\u0cbf\\u0c97\\u0cbf \\u0ca4\\u0cbe\\u0cb2\\u0cc2\\u0c95\\u0cbf\\u0ca8")
APP_THANDA = u("\\u0c85\\u0caa\\u0ccd\\u0caa\\u0cc7\\u0ca8\\u0cb9\\u0cb3\\u0ccd\\u0cb3\\u0cbf \\u0ca4\\u0cbe\\u0c82\\u0ca1")
APP_THANDAA = u("\\u0c85\\u0caa\\u0ccd\\u0caa\\u0cc7\\u0ca8\\u0cb9\\u0cb3\\u0ccd\\u0cb3\\u0cbf \\u0ca4\\u0cbe\\u0c82\\u0ca1\\u0cbe")
STATUS_KN = u("\\u0c9a\\u0cbe\\u0cb2\\u0ccd\\u0ca4\\u0cbf\\u0caf\\u0cb2\\u0ccd\\u0cb2\\u0cbf\\u0ca6\\u0cc6")
OTTU = u("\\u0c92\\u0c9f\\u0ccd\\u0c9f\\u0cc1")
KR = u("\\u0c95\\u0ccd\\u0cb0")
KAMAGARI = u("\\u0c95\\u0cbe\\u0cae\\u0c97\\u0cbe\\u0cb0\\u0cbf")

EXTRA_GPS = {
    "Sunkadakallu": {
        "name": "Sunkadakallu",
        "nameKn": SUNKA_KN,
        "villages": [{"name": "Sunkadakallu", "nameKn": SUNKA_KN}],
    },
    "Kudligi Town": {
        "name": "Kudligi Town",
        "nameKn": KUD_TOWN_KN,
        "villages": [{"name": "Kudligi", "nameKn": KUD_KN}],
    },
    "Kudligi Constituency": {
        "name": "Kudligi Constituency",
        "nameKn": KUD_CONST_KN,
        "villages": [
            {"name": "Kudligi Taluk", "nameKn": TALUK_KN},
            {"name": "Constituency", "nameKn": KUD_CONST_KN},
        ],
    },
}

VILLAGE_OVERRIDES = {
    ("Appenahalli", APP_THANDA): "Appenahalli Thanda",
    ("Appenahalli", APP_THANDAA): "Appenahalli Thanda",
    ("Kudligi Town", KUD_KN): "Kudligi",
    ("Kudligi Constituency", TALUK_KN): "Kudligi Taluk",
    ("Sunkadakallu", SUNKA_KN): "Sunkadakallu",
}


def norm(s: str) -> str:
    s = unicodedata.normalize("NFC", (s or "").strip())
    s = s.replace("\u200c", "").replace("\u200d", "")
    s = re.sub(r"\s+", " ", s)
    return s


def soft(s: str) -> str:
    """Loose KN/EN compare: drop spaces, dots, virama, common length marks."""
    s = norm(s).lower()
    s = s.replace(".", "").replace(" ", "").replace("-", "")
    # Kannada virama + length / similar vowel marks that often differ in sheets
    for ch in (
        "\u0ccd",  # virama
        "\u0cbe",  # aa
        "\u0cc0",  # ii
        "\u0cc2",  # uu
        "\u0cc7",  # ee
        "\u0ccb",  # oo
    ):
        s = s.replace(ch, "")
    return s


def find_xlsx() -> Path:
    for f in BACKEND.glob("*.xlsx"):
        if f.name == "BACKEND DATA.xlsx":
            continue
        if KAMAGARI in f.name or "2023" in f.name:
            return f
    raise FileNotFoundError("Development works xlsx not found in backend/")


def parse_gp_js() -> dict:
    text = GP_JS.read_text(encoding="utf-8")
    blocks = re.findall(
        r'\{\s*name:\s*"([^"]+)",\s*nameKn:\s*"([^"]*)",\s*villages:\s*\[(.*?)\],\s*\}',
        text,
        flags=re.S,
    )
    gps = {}
    for name, name_kn, villages_blob in blocks:
        villages = []
        for vm in re.finditer(
            r'\{\s*name:\s*"([^"]+)",\s*nameKn:\s*"([^"]*)"\s*\}',
            villages_blob,
        ):
            villages.append({"name": vm.group(1), "nameKn": vm.group(2)})
        gps[name] = {"name": name, "nameKn": name_kn, "villages": villages}
    return gps


def build_village_index(gp: dict) -> dict:
    idx = {}
    for v in gp.get("villages", []):
        idx[norm(v["nameKn"])] = v["name"]
        idx[norm(v["name"])] = v["name"]
        idx[re.sub(r"[\s.]+", "", norm(v["nameKn"]))] = v["name"]
        idx[re.sub(r"[\s.]+", "", norm(v["name"]))] = v["name"]
    return idx


def resolve_village(gp_en: str, village_kn: str, gp: dict, idx: dict):
    raw = norm(village_kn)
    if not raw:
        if gp.get("villages"):
            return gp["villages"][0]["name"], "default-first-village"
        return gp_en, "default-gp-name"
    key = (gp_en, raw)
    if key in VILLAGE_OVERRIDES:
        return VILLAGE_OVERRIDES[key], "override"
    if raw in idx:
        return idx[raw], "exact-kn"
    compact = re.sub(r"[\s.]+", "", raw)
    if compact in idx:
        return idx[compact], "compact"

    soft_raw = soft(raw)
    # Match GP HQ spelling variants → first / same-named village
    if soft_raw and (
        soft_raw == soft(gp.get("nameKn", ""))
        or soft_raw == soft(gp.get("name", ""))
        or soft_raw == soft(gp_en)
    ):
        for v in gp.get("villages", []):
            if soft(v["name"]) == soft(gp_en) or soft(v["nameKn"]) == soft_raw:
                return v["name"], "gp-hq"
        if gp.get("villages"):
            return gp["villages"][0]["name"], "gp-hq-first"

    soft_idx = {soft(k): v for k, v in idx.items() if soft(k)}
    if soft_raw in soft_idx:
        return soft_idx[soft_raw], "soft"

    for v in gp.get("villages", []):
        kn = norm(v["nameKn"])
        if not kn:
            continue
        if kn in raw or raw in kn:
            return v["name"], "contains"
        if kn.replace(" ", "") in compact or compact in kn.replace(" ", ""):
            return v["name"], "contains-compact"
        sk = soft(kn)
        if sk and (sk in soft_raw or soft_raw in sk) and min(len(sk), len(soft_raw)) >= 4:
            return v["name"], "soft-contains"
    return raw, "unmatched-keep-kn"


def parse_amount_lakhs(val) -> float:
    if val is None or val == "":
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "")
    s = re.sub(r"[^\d.]", "", s)
    if not s:
        return 0.0
    return float(s)


def join_shara(c4, c5) -> str:
    parts = []
    for c in (c4, c5):
        if c is None or c == "":
            continue
        t = str(c).strip()
        if t and t not in parts:
            parts.append(t)
    return " / ".join(parts)


def is_total_row(village: str, work: str) -> bool:
    blob = f"{village} {work}"
    return OTTU in blob or "TOTAL" in blob.upper()


def js_str(s: str) -> str:
    return json.dumps(s or "", ensure_ascii=False)


def main() -> None:
    xlsx = find_xlsx()
    gps = parse_gp_js()
    gps.update(EXTRA_GPS)

    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    developments = []
    report = {
        "file": xlsx.name,
        "sheets": [],
        "unmatchedVillages": [],
        "unknownSheets": [],
        "total": 0,
    }

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        key = norm(sheet_name)
        gp_en = SHEET_GP_MAP.get(key)
        if not gp_en:
            report["unknownSheets"].append(sheet_name)
            continue

        gp = gps.get(gp_en) or EXTRA_GPS.get(gp_en) or {
            "name": gp_en,
            "nameKn": key,
            "villages": [],
        }
        idx = build_village_index(gp)
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        header_i = 1
        for i, r in enumerate(rows[:6]):
            if not r or r[0] is None:
                continue
            a0 = str(r[0])
            if KR in a0 or a0.strip().lower() in {"si", "s.no", "sl.no", "sl"}:
                header_i = i
                break

        sheet_count = 0
        last_village_kn = ""
        for r in rows[header_i + 1 :]:
            if not r:
                continue
            village_kn = norm(str(r[1])) if r[1] is not None else ""
            work = norm(str(r[2])) if len(r) > 2 and r[2] is not None else ""
            if not work:
                continue
            if is_total_row(village_kn, work):
                continue
            if not village_kn:
                village_kn = last_village_kn
            else:
                last_village_kn = village_kn

            village_en, how = resolve_village(gp_en, village_kn, gp, idx)
            if how.startswith("unmatched"):
                report["unmatchedVillages"].append(
                    {"gp": gp_en, "villageKn": village_kn, "work": work[:80]}
                )

            amount_lakhs = parse_amount_lakhs(r[3] if len(r) > 3 else None)
            amount_inr = round(amount_lakhs * 100000, 2)
            shara = join_shara(r[4] if len(r) > 4 else None, r[5] if len(r) > 5 else None)

            department = ""
            yojane = ""
            if shara:
                parts = [p.strip() for p in re.split(r"[/,|]", shara) if p.strip()]
                if parts:
                    department = parts[0]
                if len(parts) > 1:
                    yojane = " / ".join(parts[1:])

            developments.append(
                {
                    "gramPanchayat": gp_en,
                    "village": village_en,
                    "name": work[:120],
                    "nameKn": work,
                    "description": work,
                    "descriptionKn": work,
                    "details": shara,
                    "detailsKn": shara,
                    "amountSanctioned": amount_inr,
                    "amountLakhs": amount_lakhs,
                    "status": "Ongoing",
                    "statusKn": STATUS_KN,
                    "beneficiaries": "",
                    "beneficiariesKn": "",
                    "department": department,
                    "departmentKn": department,
                    "startDate": None,
                    "locationNote": village_en,
                    "locationNoteKn": village_kn or village_en,
                    "yojane": yojane,
                    "yojaneKn": yojane,
                    "shara": shara,
                }
            )
            sheet_count += 1

        report["sheets"].append({"sheet": sheet_name, "gp": gp_en, "count": sheet_count})

    wb.close()
    report["total"] = len(developments)

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(
        json.dumps(developments, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    REPORT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    lines = [
        "/** Development works imported from constituency kamagari xlsx (2023-2026).",
        " * Regenerated by: frontend/scripts/import_developments_from_xlsx.py",
        " */",
        "export const seedDevelopments = [",
    ]
    for i, d in enumerate(developments, start=1):
        lines += [
            "  {",
            f'    id: "dev-{i:04d}",',
            f"    gramPanchayat: {js_str(d['gramPanchayat'])},",
            f"    village: {js_str(d['village'])},",
            f"    name: {js_str(d['name'])},",
            f"    nameKn: {js_str(d['nameKn'])},",
            f"    description: {js_str(d['description'])},",
            f"    descriptionKn: {js_str(d['descriptionKn'])},",
            f"    details: {js_str(d['details'])},",
            f"    detailsKn: {js_str(d['detailsKn'])},",
            f"    amountSanctioned: {d['amountSanctioned']},",
            f"    status: {js_str(d['status'])},",
            f"    statusKn: {js_str(d['statusKn'])},",
            f"    beneficiaries: {js_str(d['beneficiaries'])},",
            f"    beneficiariesKn: {js_str(d['beneficiariesKn'])},",
            f"    department: {js_str(d['department'])},",
            f"    departmentKn: {js_str(d['departmentKn'])},",
            "    startDate: null,",
            f"    locationNote: {js_str(d['locationNote'])},",
            f"    locationNoteKn: {js_str(d['locationNoteKn'])},",
            "    images: [],",
            f"    yojane: {js_str(d['yojane'])},",
            f"    yojaneKn: {js_str(d['yojaneKn'])},",
            '    updatedAt: "2026-08-01T00:00:00.000Z",',
            "  },",
        ]
    lines += ["];", ""]
    OUT_JS.write_text("\n".join(lines) + "\n", encoding="utf-8")

    print(
        json.dumps(
            {
                "total": report["total"],
                "sheets": len(report["sheets"]),
                "unknownSheets": report["unknownSheets"],
                "unmatchedVillages": len(report["unmatchedVillages"]),
                "outJson": str(OUT_JSON.relative_to(ROOT)),
                "outJs": str(OUT_JS.relative_to(ROOT)),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
